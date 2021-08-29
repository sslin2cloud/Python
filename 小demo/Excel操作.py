# coding=utf-8

# Excel操作库 xlrd
import xlrd
import xlwt
import datetime
import os
import openpyxl

# ----------------------------------- xlrd/xlwt -----------------------------------
class excel_ctrl_xl:
    def __init__(self, file) -> None:
        self.path = file
        self.wb = xlrd.open_workbook(file)
    
    def show_sheet_by_index(self, index):
        self.sh = self.wb.sheet_by_index(index)
        print("rows: " + str(self.sh.nrows))
        print("cols: " + str(self.sh.ncols))

    def show_sheet_by_name(self, name):
        self.sh = self.wb.sheet_by_name(name)
        print("rows: " + str(self.sh.nrows))
        print("cols: " + str(self.sh.ncols))

    def get_col_by_name(self, name):
        self.sh = self.wb.sheet_by_index(0)
        for i in range(self.sh.ncols):
            if self.sh.cell_value(0, i) == name:
                return i
# ----------------------------------- xlrd/xlwt -----------------------------------

# ----------------------------------- openpyxl -----------------------------------
class excel_ctrl_op:
    def __init__(self, filename) -> None:
        self.file = filename
        self.wb = openpyxl.load_workbook(filename)

    def show_sheets(self):
        print(self.wb.sheetnames)

        for sheet in self.wb:
            print(sheet.title)
    
    def show_sheet(self):
        ws = self.wb['Sheet1']
        print(ws.max_row)
        print(ws.max_column)

        print(ws['A1'])
        print(ws.cell(row=1, column=2))
        print(ws['A1'].value)      #读取单元格的值
        print(ws['A1'].row)        #读取的表格的行数
        print(ws['A1'].column)     #读取的表格的列数
        print(ws['A1'].coordinate) #读取的表格的行列数。输出的值为'A1'
# ----------------------------------- openpyxl -----------------------------------

c = excel_ctrl_xl(r"D:/Github/demo.xlsx")
c.show_sheet_by_index(0)
print(c.get_col_by_name('a'))

c = excel_ctrl_op(r"D:/Github/demo.xlsx")
c.show_sheets()
c.show_sheet()